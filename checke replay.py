import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

st.set_page_config(layout="wide")
st.title("☑️ Oversent Verification Tool ")

# =========================
# SESSION STATE INIT
# =========================
if "results_df" not in st.session_state:
    st.session_state.results_df = None

# =========================
# RESET FUNCTION
# =========================
def reset_app():
    st.session_state.results_df = None
    st.session_state.uploaded_files = None
    st.session_state.model = ""
    st.session_state.odf = ""
    st.rerun()

# =========================
# UPLOAD FILES
# =========================
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = None

files = st.file_uploader(
    "📂 Upload FRS Files",
    type=["xlsx"],
    accept_multiple_files=True,
    key="file_uploader"
)

if files:
    st.session_state.uploaded_files = files

file_names = [f.name for f in files] if files else []

# =========================
# GLOBAL INPUT
# =========================
model = st.text_input("📌 Model", key="model")
odf = st.text_input("📌 ODF", key="odf").strip().upper()

selected_file = st.selectbox("📂 Select FRS File", file_names)

# =========================
# TABLE SIZE
# =========================
st.subheader("🧾 Create Your Table")

num_rows = st.number_input("Number of articles", min_value=1, value=5)

df_input = pd.DataFrame({
    "PART NO": [""] * num_rows,
    "QTY NEEDED": [0] * num_rows,
    "QTY SENT": [0] * num_rows,
    "OVERSENT REPLY": [0] * num_rows
})

edited_df = st.data_editor(df_input, use_container_width=True)

# =========================
# PROCESS
# =========================
if st.button("▶️ Calculate"):

    if not files:
        st.error("❌ Upload file first")
        st.stop()

    file_obj = next(f for f in files if f.name == selected_file)

    df = pd.read_excel(file_obj)
    df.columns = df.columns.str.strip().str.upper()

    part_col = next((c for c in df.columns if "PART" in c), None)
    odf_col = next((c for c in df.columns if "ODF" in c), None)
    oversent_col = next((c for c in df.columns if "OVERSENT QTY" in c), None)

    if oversent_col is None:
        st.error("❌ Column 'OVERSENT QTY' not found")
        st.stop()

    df[part_col] = df[part_col].astype(str).str.strip().str.upper()
    df[odf_col] = df[odf_col].astype(str).str.strip().str.upper()

    df = df.reset_index(drop=True)

    results = []

    for _, row in edited_df.iterrows():

        pn = str(row["PART NO"]).strip().upper()
        qty_needed = row["QTY NEEDED"]
        qty_sent = row["QTY SENT"]
        oversent_reply = row["OVERSENT REPLY"]

        if pn == "":
            continue

        result = {
            "MODEL": model,
            "PART NO": pn,
            "LAST OVERSENT": None,
            "QTY NEEDED": qty_needed,
            "QTY SENT": qty_sent,
            "CALCULATED OVERSENT": None,
            "OVERSENT REPLY": oversent_reply,
            "STATUS": ""
        }

        same_pn = df[df[part_col] == pn]

        if same_pn.empty:
            result["STATUS"] = "PN NOT FOUND"
            results.append(result)
            continue

        matches_odf = same_pn[same_pn[odf_col] == odf]

        if matches_odf.empty:
            result["STATUS"] = "ODF NOT FOUND"
            results.append(result)
            continue

        current_idx = matches_odf.index[0]

        previous_rows = same_pn[same_pn.index < current_idx]

        if not previous_rows.empty:
            last_row = previous_rows.iloc[-1]
            last_oversent = last_row[oversent_col]
        else:
            last_oversent = 0

        result["LAST OVERSENT"] = last_oversent

        calc = (last_oversent - qty_needed) + qty_sent
        result["CALCULATED OVERSENT"] = calc

        result["STATUS"] = "OK" if calc == oversent_reply else "NON CONFORME"

        results.append(result)

    st.session_state.results_df = pd.DataFrame(results)

    st.success("✅ Calculation Done")

# =========================
# DISPLAY + EXPORT
# =========================
if st.session_state.results_df is not None:

    df_result = st.session_state.results_df

    def color_status(val):
        if val == "OK":
            return "background-color: #c6f7c6"
        elif val == "NON CONFORME":
            return "background-color: #f7c6c6"
        return ""

    styled_df = df_result.style.map(color_status, subset=["STATUS"])

    st.dataframe(styled_df, use_container_width=True)

    # =========================
    # EXPORT WITH HEADER + BORDER FIX
    # =========================
    def export_excel(df):
        wb = Workbook()
        ws = wb.active

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        green = PatternFill(start_color="C6F7C6", end_color="C6F7C6", fill_type="solid")
        red = PatternFill(start_color="F7C6C6", end_color="F7C6C6", fill_type="solid")

        # HEADER
        headers = list(df.columns)
        ws.append(headers)

        # 🔥 APPLY BORDER TO HEADER
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.border = thin_border

        # DATA
        for row in df.itertuples(index=False):
            ws.append(list(row))

        # APPLY STYLE (ALL CELLS + HEADER BORDER FIX)
        for i, row in enumerate(df.itertuples(), start=2):

            status = row.STATUS

            fill = green if status == "OK" else red

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=i, column=col)
                cell.fill = fill
                cell.border = thin_border

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    output = export_excel(df_result)

    st.download_button(
        "📥 Download Result",
        data=output,
        file_name="Oversent_Result.xlsx"
    )

# =========================
# RESET BUTTON
# =========================
st.button("🧹 Reset Data", on_click=reset_app)
